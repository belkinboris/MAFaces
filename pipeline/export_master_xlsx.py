#!/usr/bin/env python3
"""Сводная таблица всех собранных сделок в один Excel-файл с одинаковыми столбцами.

ВАЖНО: источники данных должны быть открытыми (публичные новости, сайты юрфирм,
телеграм-каналы) — НЕ платные структурированные базы вроде mergers.ru/deals
(там прямо указано "детальная информация предоставляется только на коммерческой
основе") или СПАРК-Интерфакс. Этот скрипт сводит воедино то, что мы УЖЕ легально
собрали через collect_deals.py / enrich_deals.py.

Столбцы (одинаковые для всех сделок, независимо от источника):
  date | title | industry | amount_text | amount_value | amount_unit | amount_currency
  | amount_rub_approx_mln | status | firm | source_name | source_url | is_duplicate_of

Использование:
  python3 export_master_xlsx.py in1.json in2.json ... --out master_deals.xlsx

Каждый входной файл — список записей (как отдаёт collect_deals.py/enrich_deals.py
после --dedupe-only). Скрипт сам сольёт дубли по (год ± 30 дней + пересечение
заголовков) и отсортирует по приблизительной сумме в рублях (только для ранжирования,
не как факт — курс USD/RUB фиксированный и грубый, см. столбец amount_rub_approx_mln).
"""
import argparse
import json
import re
import sys

USD_RUB = 90  # грубый фиксированный курс ТОЛЬКО для сортировки между валютами, не факт

UNIT_MULT = {"млрд": 1000, "млн": 1, "тыс": 0.001}

ESTIMATE_WORDS = re.compile(
    r"оценк|аналит|эксперт|возможн|стартов|начальн|по некоторым данным|не раскрыв|"
    r"предполож|ориентировочн|официально не раскрывается", re.I)


EXCLUDE_CONTEXT = re.compile(
    r"иск\w*|штраф\w*|ущерб\w*|неустойк\w*|задолженност\w*|долг\w*(?!ов[аи] сделк)|"
    r"выручк\w*|капитализац\w*|оборот\w*|инвестиц\w* в развит", re.I)


def parse_amount(text):
    """Возвращает (значение, единица, валюта, is_official).
    Если сумма — оценка аналитика/не раскрыта официально, ЛИБО число стоит рядом
    со словом, не относящимся к цене сделки (иск, штраф, выручка, капитализация) —
    число НЕ извлекается (is_official=False). Свободный текст может содержать
    несколько разных чисел про разное — берём только то, что похоже на цену сделки."""
    if not text or "не раскрыт" in text.lower():
        return None, None, None, False
    is_official = not ESTIMATE_WORDS.search(text)
    t = text.replace(",", ".")
    m = re.search(r"([\d.]+)\s*[-–—]\s*[\d.]+\s*(млрд|млн|тыс)", t, re.I)
    if not m:
        m = re.search(r"([\d.]+)\s*(млрд|млн|тыс)", t, re.I)
    if not m:
        return None, None, None, False
    # смотрим 60 символов до найденного числа — если там слово из чужого контекста, не доверяем
    window = t[max(0, m.start() - 60):m.start()]
    if EXCLUDE_CONTEXT.search(window):
        return None, None, None, False
    val = float(m.group(1))
    unit = m.group(2).lower()
    local = t[max(0, m.start() - 15):m.end() + 15]  # валюту смотрим ТОЛЬКО рядом с самим числом
    currency = "USD" if ("$" in local or "долл" in local.lower()) else ("EUR" if ("€" in local or "евро" in local.lower()) else "RUB")
    return val, unit, currency, is_official


def to_rub_millions(val, unit, currency):
    if val is None:
        return None
    base = val * UNIT_MULT.get(unit, 1)  # -> млн в исходной валюте
    if currency == "USD":
        return round(base * USD_RUB, 1)
    if currency == "EUR":
        return round(base * USD_RUB * 1.05, 1)  # грубая оценка EUR/USD, только для сортировки
    return round(base, 1)


def norm_date(d):
    d = (d or "").strip()
    if len(d) == 4:
        return d + "-06-15"
    if len(d) == 7:
        return d + "-15"
    return d[:10] if len(d) >= 10 else "unknown"


STOP = {"ооо", "ао", "пао", "гк", "компания", "группа", "доля", "долей", "акций", "сделка",
        "бизнес", "приобрел", "приобрела", "купил", "купила", "может", "купить", "продал",
        "продала", "россии", "покупает", "инвестирует"}


def toks(t):
    return {w for w in re.sub(r"[«»\"'().,:;–—-]", " ", (t or "").lower()).split()
            if len(w) > 3 and w not in STOP}


def dedupe(records):
    """Помечает дубли (не удаляет), чтобы в Excel было видно, что смержено."""
    for r in records:
        r.setdefault("_dup_of", None)
    for i, a in enumerate(records):
        if a["_dup_of"] is not None:
            continue
        da = norm_date(a.get("date", ""))[:4]
        ta = toks(a.get("title", ""))
        for j in range(i + 1, len(records)):
            b = records[j]
            if b["_dup_of"] is not None:
                continue
            db = norm_date(b.get("date", ""))[:4]
            tb = toks(b.get("title", ""))
            if da == db and ta and tb and len(ta & tb) >= 3:
                b["_dup_of"] = a.get("title", "")[:60]
    return records


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs", nargs="+", help="JSON-файлы со сделками (списки записей)")
    ap.add_argument("--out", default="master_deals.xlsx")
    args = ap.parse_args()

    all_records = []
    for path in args.inputs:
        try:
            data = json.load(open(path, encoding="utf-8"))
        except FileNotFoundError:
            print(f"  ! не найден: {path}", file=sys.stderr)
            continue
        print(f"  {path}: {len(data)} записей")
        all_records += data

    print(f"Всего до дедупа: {len(all_records)}")
    all_records = dedupe(all_records)
    unique = sum(1 for r in all_records if r["_dup_of"] is None)
    print(f"Уникальных (не помечено дублем): {unique}")

    rows = []
    for r in all_records:
        sum_text = r.get("sum") or r.get("role") or ""
        val, unit, cur, is_official = parse_amount(sum_text)
        rub_approx = to_rub_millions(val, unit, cur) if is_official else None
        rows.append({
            "date": norm_date(r.get("date", "")),
            "title": r.get("title", ""),
            "industry": r.get("industry") or r.get("ind") or "",
            "amount_text": (sum_text[:200] if sum_text else "Не раскрыта"),
            "amount_value": val if is_official else None,
            "amount_unit": unit or "" if is_official else "",
            "amount_currency": cur or "" if is_official else "",
            "amount_rub_approx_mln": rub_approx,
            "amount_is_official": "Официально" if is_official and val else ("Оценка/не раскрыта" if val else ""),
            "status": r.get("status") or r.get("confidence") or "",
            "firm": r.get("firm_id") or r.get("firm") or "",
            "source_name": r.get("source_name") or (r.get("src", [None])[0] if r.get("src") else ""),
            "source_url": r.get("source_url") or (r.get("src", [None, None])[1] if r.get("src") else ""),
            "is_duplicate_of": r["_dup_of"] or "",
        })

    # сортировка: сначала официально раскрытые суммы (по убыванию), потом всё остальное
    rows.sort(key=lambda x: (x["amount_rub_approx_mln"] is None, -(x["amount_rub_approx_mln"] or 0)))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"
    headers = ["Дата", "Название", "Отрасль", "Сумма (текст, до 200 симв.)", "Значение", "Ед.", "Валюта",
               "≈ млн ₽ (для сортировки)", "Статус суммы", "Стадия/уверенность", "Фирма", "Источник", "URL", "Дубль записи"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F0EFEA")

    for r in rows:
        ws.append([r["date"], r["title"], r["industry"], r["amount_text"], r["amount_value"],
                   r["amount_unit"], r["amount_currency"], r["amount_rub_approx_mln"], r["amount_is_official"],
                   r["status"], r["firm"], r["source_name"], r["source_url"], r["is_duplicate_of"]])

    widths = [12, 55, 18, 42, 10, 8, 8, 20, 16, 14, 14, 16, 40, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # примечание про приблизительность курса
    note_row = len(rows) + 3
    ws.cell(row=note_row, column=1,
            value="«≈ млн ₽» — грубая оценка по фиксированному курсу 90 ₽/$, ТОЛЬКО для сортировки. "
                  "«Статус суммы»=«Оценка/не раскрыта» — сумма НЕ учтена в сортировке (это догадка "
                  "аналитика или прямо помечено как нераскрытое), но текст показан для справки.")

    wb.save(args.out)
    print(f"\n→ {args.out}: {len(rows)} строк")
    with_sum = sum(1 for r in rows if r["amount_rub_approx_mln"])
    print(f"С распознанной суммой (можно сортировать по цене): {with_sum}")


if __name__ == "__main__":
    main()
