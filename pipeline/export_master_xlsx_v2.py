#!/usr/bin/env python3
"""Свод харвест-файлов (mergers/firms/tg-firms/tg-news) в Excel для отбора топ-N.

Честность: суммы-оценки/иски/выручка не считаются раскрытой ценой сделки.
Дедуп: помечает записи, которые уже обогащены (enriched_2026_final.json) или
уже стали полными карточками сайта (deals_2026.json / index.html DEALS) —
чтобы не тратить бюджет повторно.
"""
import json
import re
import sys

ESTIMATE_WORDS = re.compile(
    r"оценк|аналит|эксперт|возможн|стартов|начальн|по некоторым данным|не раскрыв|"
    r"предполож|ориентировочн|официально не раскрывается|могл[аи]?\s|не отказал|"
    r"рассматрива", re.I)
NOT_A_DEAL = re.compile(
    r"конгресс|суд[а-я]* (поддержал|отклонил|обязал|взыскал)|по всему миру|"
    r"комментарий в связи|статья [а-я]* об|анонсировал[аи]? визит|"
    r"компенсаци\w*|исковое заявлени", re.I)
EXCLUDE_CONTEXT = re.compile(
    r"иск\w*|штраф\w*|ущерб\w*|неустойк\w*|задолженност\w*|"
    r"выручк\w*|капитализац\w*|оборот\w*|инвестиц\w* в развит", re.I)
UNIT_MULT = {"млрд": 1000, "млн": 1, "тыс": 0.001}
USD_RUB = 90

STOP = {"ооо", "ао", "пао", "гк", "компания", "группа", "доля", "долей", "акций", "сделка",
        "бизнес", "приобрел", "приобрела", "купил", "купила", "может", "купить", "продал",
        "продала", "россии", "покупает", "инвестирует", "выкупит", "выкупил", "получил"}


def toks(t):
    return {w for w in re.sub(r"[«»\"'().,:;–—-]", " ", (t or "").lower()).split()
            if len(w) > 3 and w not in STOP}


def parse_amount(text_for_amount, full_context, title=""):
    """text_for_amount — короткая строка с суммой (из заголовка или найденная в тексте).
    full_context — весь доступный текст, чтобы проверить контекст вокруг числа."""
    if not text_for_amount:
        return None, None, None, False
    if NOT_A_DEAL.search(title or ""):
        return None, None, None, False
    t = text_for_amount.replace(",", ".")
    is_official = not ESTIMATE_WORDS.search((full_context or "") + " " + (title or "") + " " + text_for_amount)
    m = re.search(r"(\d[\d.]*)\s*[-–—]\s*\d[\d.]*\s*(млрд|млн|тыс)", t, re.I)
    if not m:
        m = re.search(r"(\d[\d.]*)\s*(млрд|млн|тыс)", t, re.I)
    if not m:
        return None, None, None, False
    try:
        val = float(m.group(1))
    except ValueError:
        return None, None, None, False
    # контекст вокруг найденного числа в ПОЛНОМ тексте (не только в коротком фрагменте)
    ctx = full_context or t
    idx = ctx.find(m.group(0))
    window = ctx[max(0, idx - 60):idx] if idx >= 0 else ""
    if EXCLUDE_CONTEXT.search(window):
        return None, None, None, False
    unit = m.group(2).lower()
    local = t[max(0, m.start() - 15):m.end() + 15]
    currency = "USD" if ("$" in local or "долл" in local.lower()) else ("EUR" if ("€" in local or "евро" in local.lower()) else "RUB")
    return val, unit, currency, is_official


def to_rub_millions(val, unit, currency):
    if val is None:
        return None
    base = val * UNIT_MULT.get(unit, 1)
    if currency == "USD":
        return round(base * USD_RUB, 1)
    if currency == "EUR":
        return round(base * USD_RUB * 1.05, 1)
    return round(base, 1)


def norm_date(d):
    d = (d or "").strip()
    return d[:10] if len(d) >= 10 else "unknown"


def load_already_done():
    """Собирает заголовки уже обогащённых/готовых сделок для исключения из кандидатов."""
    done_titles = []
    try:
        enr = json.load(open("/mnt/user-data/uploads/enriched_2026_final.json", encoding="utf-8"))
        done_titles += [r.get("title", "") for r in enr if r.get("enrichment_confidence") in ("high", "medium")]
    except FileNotFoundError:
        pass
    try:
        d26 = json.load(open("/home/claude/reestr/static/data/deals_2026.json", encoding="utf-8"))
        done_titles += [r.get("title", "") for r in d26.get("deals", [])]
    except FileNotFoundError:
        pass
    # 17 ручных карточек — вытаскиваем заголовки прямо из index.html
    try:
        html = open("/home/claude/reestr/static/index.html", encoding="utf-8").read()
        m = re.search(r"let DEALS = \[([\s\S]*?)\n\];", html)
        if m:
            done_titles += re.findall(r'title:"([^"]+)"', m.group(1))
    except FileNotFoundError:
        pass
    return done_titles


def main():
    files = {
        "mergers": "/mnt/user-data/uploads/collected_mergers_titles.json",
        "firms": "/mnt/user-data/uploads/collected_firms_titles.json",
        "tg_firms": "/mnt/user-data/uploads/collected_tg_firms_titles.json",
        "tg_news": "/mnt/user-data/uploads/collected_tg_news_titles.json",
    }
    all_records = []
    for src, path in files.items():
        try:
            data = json.load(open(path, encoding="utf-8"))
        except FileNotFoundError:
            print(f"  ! не найден: {path}", file=sys.stderr)
            continue
        print(f"  {src}: {len(data)} записей")
        for r in data:
            r["_src_file"] = src
        all_records += data
    print(f"Всего сырых записей: {len(all_records)}")

    done_titles = load_already_done()
    done_toks = [toks(t) for t in done_titles if t]
    print(f"Уже обогащено/готово (для исключения): {len(done_toks)} записей")

    # дедуп внутри собранного (год + пересечение токенов >=3)
    for r in all_records:
        r["_dup_of"] = None
        r["_already_done"] = False
    by_year = {}
    for i, r in enumerate(all_records):
        y = norm_date(r.get("date", ""))[:4]
        by_year.setdefault(y, []).append(i)

    for y, idxs in by_year.items():
        for pos, i in enumerate(idxs):
            if all_records[i]["_dup_of"] is not None:
                continue
            ti = toks(all_records[i].get("title", ""))
            if not ti:
                continue
            for j in idxs[pos + 1:]:
                if all_records[j]["_dup_of"] is not None:
                    continue
                tj = toks(all_records[j].get("title", ""))
                if tj and len(ti & tj) >= 3:
                    all_records[j]["_dup_of"] = all_records[i].get("title", "")[:60]

    already_count = 0
    for r in all_records:
        rt = toks(r.get("title", ""))
        if not rt:
            continue
        for dt in done_toks:
            if len(rt & dt) >= 3:
                r["_already_done"] = True
                already_count += 1
                break
    print(f"Совпало с уже готовым (будет исключено из кандидатов): {already_count}")

    rows = []
    for r in all_records:
        amt_text = r.get("amount_from_title") or ""
        full_text = r.get("raw_text") or amt_text
        val, unit, cur, is_official = parse_amount(amt_text or full_text, full_text, r.get("title", ""))
        rub = to_rub_millions(val, unit, cur) if is_official else None
        rows.append({
            "date": norm_date(r.get("date", "")),
            "title": r.get("title", ""),
            "amount_text": amt_text or (full_text[:150] if full_text else ""),
            "amount_rub_mln": rub,
            "is_official": "Официально" if (is_official and val) else ("Не подтверждено" if val else ""),
            "source": r.get("source", r.get("_src_file", "")),
            "firm": r.get("firm_id", ""),
            "url": r.get("url", ""),
            "is_duplicate_of": r["_dup_of"] or "",
            "already_enriched": "ДА — пропустить" if r["_already_done"] else "",
        })

    rows.sort(key=lambda x: (x["amount_rub_mln"] is None, -(x["amount_rub_mln"] or 0)))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"
    headers = ["Дата", "Название", "Сумма (текст)", "≈ млн ₽", "Статус суммы", "Источник",
               "Фирма", "URL", "Дубль записи", "Уже обогащено"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F0EFEA")
    for r in rows:
        ws.append([r["date"], r["title"], r["amount_text"][:150], r["amount_rub_mln"],
                   r["is_official"], r["source"], r["firm"], r["url"], r["is_duplicate_of"],
                   r["already_enriched"]])
    widths = [12, 60, 40, 14, 16, 12, 12, 45, 30, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save("/tmp/master_deals_v2.xlsx")

    # сводка для бюджетного разговора
    candidates = [r for r in rows if r["amount_rub_mln"] and not r["already_done" if False else "already_enriched"]]
    fresh_candidates = [r for r in rows if r["amount_rub_mln"] and not r["already_enriched"] and not r["is_duplicate_of"]]
    print(f"\n=== СВОДКА ===")
    print(f"Всего строк в Excel: {len(rows)}")
    print(f"С официальной суммой: {sum(1 for r in rows if r['amount_rub_mln'])}")
    print(f"С суммой И не дубль И ещё не обогащено (реальные кандидаты): {len(fresh_candidates)}")


if __name__ == "__main__":
    main()
